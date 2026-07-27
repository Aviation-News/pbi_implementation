from flask import Flask, render_template, jsonify, request, session, redirect, url_for, Response
from flask_cors import CORS
import requests
import json
import os
import time
import uuid
import csv
import io
import re
import openpyxl
from openpyxl.styles import Font

app = Flask(__name__)
CORS(app, supports_credentials=True)  # Fix CORS issues

EXPORT_ROW_LIMIT = 3000  # change here to update the cap everywhere
app.secret_key = os.environ.get('PBI_SECRET_KEY', 'change-me-pbi-saved-filters')  # stable key so sessions survive restart

# Default configuration
DEFAULT_CONFIG = {
  'tenant_id': '65693725-b6ed-4482-838d-454f6f79eafb',
  'client_id': '8264d2da-a712-455b-bb14-6e10d4db8676',
  'client_secret': 'gH68Q~DVK4FHki2hwdgOC4we8tF3nZMsM50FhcC9',
  'group_id': '8578855e-dbb3-47b0-b302-437d8572a9a3',
  'report_id': 'd32c0ba1-80fc-4e7a-99f6-f72d42ec61b7'
}

# Where saved filters live (JSON file next to this app)
FILTERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'saved_filters.json')

# Key used to hold any pre-existing single-report saves so they aren't lost
# when we move to per-report bucketing.
LEGACY_BUCKET = '_legacy_unscoped'


def get_config():
    """Get the current configuration, either from session or default"""
    if 'config' in session:
        return session['config']
    return DEFAULT_CONFIG


def current_report_key():
    """Return the per-report bucket key for the current config.

    Saved filters are keyed by user + report so switching DEFAULT_CONFIG (or
    the in-session override) automatically gives a fresh, isolated list of
    saved sets per report.
    """
    cfg = get_config()
    return f"{cfg['group_id']}::{cfg['report_id']}"


def _load_all_filters():
    """Read the whole saved-filters store.

    Returns the new shape: {user_id: {report_key: [filter_records]}}.
    If the file is in the old shape (a list directly under the user), it is
    migrated in-memory under a LEGACY_BUCKET key so historical sets are kept
    but don't leak into any specific report's list.
    """
    if not os.path.exists(FILTERS_FILE):
        return {}
    try:
        with open(FILTERS_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}

    if not isinstance(data, dict):
        return {}

    migrated = False
    for user_id, value in list(data.items()):
        # Old shape: value is a flat list of saved sets.
        if isinstance(value, list):
            data[user_id] = {LEGACY_BUCKET: value}
            migrated = True
        elif not isinstance(value, dict):
            # Junk row; drop it.
            data.pop(user_id, None)
            migrated = True

    return data


def _save_all_filters(data):
    """Write the whole store back to disk atomically-ish."""
    tmp = FILTERS_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, FILTERS_FILE)


def get_user_id():
    """The current user key (their name/email). None if not set."""
    return session.get('user_id')


def get_access_token():
    """Get Azure AD access token for Power BI API"""
    config = get_config()
    token_url = f"https://login.microsoftonline.com/{config['tenant_id']}/oauth2/v2.0/token"
    token_data = {
        'grant_type': 'client_credentials',
        'client_id': config['client_id'],
        'client_secret': config['client_secret'],
        'scope': 'https://analysis.windows.net/powerbi/api/.default'
    }

    response = requests.post(token_url, data=token_data)
    if response.status_code == 200:
        return response.json()['access_token']
    else:
        raise Exception(f"Failed to get token: {response.text}")


def get_embed_token():
    """Get Power BI embed token and report details"""
    config = get_config()
    access_token = get_access_token()
    embed_url = f"https://api.powerbi.com/v1.0/myorg/groups/{config['group_id']}/reports/{config['report_id']}/GenerateToken"

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    embed_data = {
        'accessLevel': 'View'
    }

    response = requests.post(embed_url, headers=headers, json=embed_data)

    if response.status_code == 200:
        embed_response = response.json()
        report_url = f"https://api.powerbi.com/v1.0/myorg/groups/{config['group_id']}/reports/{config['report_id']}"
        report_response = requests.get(report_url, headers=headers)

        if report_response.status_code == 200:
            report_data = report_response.json()
            return {
                'embed_token': embed_response['token'],
                'embed_url': report_data['embedUrl'],
                'report_id': config['report_id'],
                'group_id': config['group_id']
            }
        else:
            raise Exception(f"Failed to get report data: {report_response.text}")
    else:
        raise Exception(f"Failed to get embed token: {response.text}")


@app.route('/')
def index():
    """Main page with embedded Power BI report"""
    try:
        embed_data = get_embed_token()
        return render_template(
            'index.html',
            embed_token=embed_data['embed_token'],
            embed_url=embed_data['embed_url'],
            report_id=embed_data['report_id'],
            group_id=embed_data['group_id'],
            user_id=get_user_id() or ''
        )
    except Exception as e:
        return f"""
        <div style="padding: 20px; background: #f8d7da; color: #721c24; border-radius: 5px; margin: 20px;">
            <h2>Server Error</h2>
            <p><strong>Error:</strong> {str(e)}</p>
            <p><strong>Type:</strong> {type(e).__name__}</p>
            <hr>
            <p><a href="/test" style="color: #0078d4;">Try Test Endpoint</a></p>
            <p><a href="/reset" style="color: #0078d4;">Reset Configuration</a></p>
        </div>
        """


@app.route('/update_config', methods=['POST'])
def update_config():
    """Update the configuration with user-provided values"""
    try:
        config = get_config()
        if request.form.get('workspace_id'):
            config['group_id'] = request.form.get('workspace_id')
        if request.form.get('report_id'):
            config['report_id'] = request.form.get('report_id')
        session['config'] = config
        return redirect(url_for('index'))
    except Exception as e:
        return f"""
        <div style="padding: 20px; background: #f8d7da; color: #721c24; border-radius: 5px; margin: 20px;">
            <h2>Configuration Error</h2>
            <p><strong>Error:</strong> {str(e)}</p>
            <p><a href="/" style="color: #0078d4;">Back to Main Page</a></p>
            <p><a href="/reset" style="color: #0078d4;">Reset Configuration</a></p>
        </div>
        """


@app.route('/reset')
def reset_config():
    """Reset configuration to defaults"""
    if 'config' in session:
        session.pop('config')
    return redirect(url_for('index'))


@app.route('/test')
def test():
    """Test endpoint to verify token generation"""
    try:
        embed_data = get_embed_token()
        config = get_config()
        return jsonify({
            'success': True,
            'message': 'Embed token generated successfully!',
            'data': {
                'token_length': len(embed_data['embed_token']),
                'embed_url': embed_data['embed_url'],
                'report_id': embed_data['report_id'],
                'token_preview': embed_data['embed_token'][:50] + '...'
            },
            'config': {
                'tenant_id': config['tenant_id'],
                'client_id': config['client_id'],
                'group_id': config['group_id'],
                'report_id': config['report_id']
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__
        }), 500


@app.route('/health')
def health():
    """Health check endpoint"""
    config = get_config()
    return jsonify({
        'status': 'healthy',
        'app': 'Power BI Embedded Demo',
        'port': 8080,
        'current_config': {
            'tenant_id': config['tenant_id'],
            'group_id': config['group_id'],
            'report_id': config['report_id']
        },
        'report_bucket': current_report_key(),
        'endpoints': {
            'main': 'http://localhost:8080/',
            'test': 'http://localhost:8080/test',
            'health': 'http://localhost:8080/health',
            'reset': 'http://localhost:8080/reset'
        }
    })


# ---------------------------------------------------------------------------
# User identity (simple name/email box, no real auth)
# ---------------------------------------------------------------------------
@app.route('/login', methods=['POST'])
def login():
    """Set the current user from a name/email box."""
    if request.is_json:
        user_id = (request.json or {}).get('user_id')
    else:
        user_id = request.form.get('user_id')
    user_id = (user_id or '').strip()
    if not user_id:
        return jsonify({'success': False, 'error': 'Name or email required'}), 400
    session['user_id'] = user_id
    if request.is_json:
        return jsonify({'success': True, 'user_id': user_id})
    return redirect(url_for('index'))


@app.route('/logout', methods=['POST', 'GET'])
def logout():
    """Clear the current user."""
    session.pop('user_id', None)
    if request.method == 'GET':
        return redirect(url_for('index'))
    return jsonify({'success': True})


@app.route('/whoami')
def whoami():
    return jsonify({'user_id': get_user_id(), 'report_bucket': current_report_key()})


# ---------------------------------------------------------------------------
# Saved filters API (keyed by user + report, stored in JSON file)
# ---------------------------------------------------------------------------
@app.route('/api/filters', methods=['GET'])
def list_filters():
    """List saved filter sets for the current user, scoped to the current report."""
    user_id = get_user_id()
    if not user_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    store = _load_all_filters()
    user_buckets = store.get(user_id, {})
    sets = user_buckets.get(current_report_key(), [])
    return jsonify({
        'success': True,
        'filters': sets,
        'report_bucket': current_report_key()
    })


@app.route('/api/filters', methods=['POST'])
def save_filters():
    """Save a named set of filters for the current user under the current report.

    Expected JSON body:
      { "name": "My view",
        "filters": { "filters": [...report filters...],
                     "slicers": [...slicer states...] } }
    (A bare list is also accepted for backward compatibility.)
    """
    user_id = get_user_id()
    if not user_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401

    body = request.get_json(silent=True) or {}
    name = (body.get('name') or '').strip()
    filters = body.get('filters')
    if not name:
        return jsonify({'success': False, 'error': 'Name required'}), 400
    if not isinstance(filters, (list, dict)):
        return jsonify({'success': False, 'error': 'filters must be a list or object'}), 400

    store = _load_all_filters()
    user_buckets = store.setdefault(user_id, {})
    report_key = current_report_key()
    sets_for_report = user_buckets.get(report_key, [])

    record = {
        'id': uuid.uuid4().hex,
        'name': name,
        'filters': filters,
        'saved_at': int(time.time()),
        'report_bucket': report_key
    }
    # Overwrite any existing set with the same name (user confirmed on the frontend).
    sets_for_report = [s for s in sets_for_report if s.get('name') != name]
    sets_for_report.append(record)
    user_buckets[report_key] = sets_for_report
    store[user_id] = user_buckets
    _save_all_filters(store)
    return jsonify({'success': True, 'filter': record, 'report_bucket': report_key})


@app.route('/api/filters/<filter_id>', methods=['DELETE'])
def delete_filter(filter_id):
    """Delete one saved filter set by id, scoped to the current user + report."""
    user_id = get_user_id()
    if not user_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    store = _load_all_filters()
    user_buckets = store.get(user_id, {})
    report_key = current_report_key()
    sets_for_report = user_buckets.get(report_key, [])
    new_sets = [s for s in sets_for_report if s.get('id') != filter_id]
    user_buckets[report_key] = new_sets
    store[user_id] = user_buckets
    _save_all_filters(store)
    return jsonify({'success': True, 'removed': len(sets_for_report) - len(new_sets)})


def _get_dataset_id(access_token):
    config = get_config()
    url = (f"https://api.powerbi.com/v1.0/myorg/groups/{config['group_id']}"
           f"/reports/{config['report_id']}")
    r = requests.get(url, headers={'Authorization': f'Bearer {access_token}'})
    if r.status_code != 200:
        raise Exception(f"Could not fetch report info: {r.text[:200]}")
    did = r.json().get('datasetId')
    if not did:
        raise Exception("Dataset ID not found in report info")
    return did


# Skip these common dimension/calendar table names when auto-detecting the
# primary data table from slicer filter targets.
_SKIP_TABLES = {'date', 'calendar', 'dates', 'dim_date', 'time', 'period'}


def _extract_table_name(report_filters, slicer_states):
    """Best-effort: find the dataset table name from filter/slicer targets."""
    candidates = []
    for f in (report_filters or []):
        t = (f.get('target') or {}).get('table', '')
        if t and not t.startswith('$'):
            candidates.append(t)
    for s in (slicer_states or []):
        for sf in ((s.get('state') or {}).get('filters') or []):
            t = (sf.get('target') or {}).get('table', '')
            if t and not t.startswith('$'):
                candidates.append(t)
    # Prefer any table whose name is not a pure dimension/date table
    for t in candidates:
        if t.lower() not in _SKIP_TABLES:
            return t
    return candidates[0] if candidates else None


def _filter_to_dax(f):
    """Convert one Power BI filter object to a DAX boolean expression, or None."""
    target = f.get('target') or {}
    table  = target.get('table', '')
    column = target.get('column', '')
    if not table or not column:
        return None

    col_ref = f"'{table}'[{column}]"

    # Basic filter — values list (In / NotIn)
    if 'values' in f:
        vals = [v for v in (f.get('values') or []) if v is not None]
        if not vals:
            return None
        dax_vals = ', '.join(
            f'"{v}"' if isinstance(v, str) else str(v) for v in vals
        )
        prefix = 'NOT ' if f.get('operator') == 'NotIn' else ''
        return f"{prefix}{col_ref} IN {{{dax_vals}}}"

    # Advanced filter — conditions array
    op_map = {
        'GreaterThan': '>', 'GreaterThanOrEqual': '>=',
        'LessThan': '<',    'LessThanOrEqual': '<=',
        'Is': '=',          'IsNot': '<>',
    }
    parts = []
    for cond in (f.get('conditions') or []):
        op  = op_map.get(cond.get('operator', ''))
        val = cond.get('value')
        if op and val is not None:
            dax_val = f'"{val}"' if isinstance(val, str) else str(val)
            parts.append(f"{col_ref} {op} {dax_val}")
    if parts:
        logic = ' && ' if (f.get('operator') or 'And').lower() == 'and' else ' || '
        return '(' + logic.join(parts) + ')'

    return None


def _build_dax(table, report_filters, slicer_states, row_limit=None):
    row_limit = row_limit if row_limit is not None else EXPORT_ROW_LIMIT
    safe = table.replace("'", "''")
    conditions = []
    for f in (report_filters or []):
        c = _filter_to_dax(f)
        if c:
            conditions.append(c)
    for s in (slicer_states or []):
        for sf in ((s.get('state') or {}).get('filters') or []):
            c = _filter_to_dax(sf)
            if c:
                conditions.append(c)
    if conditions:
        cond_str = ',\n    '.join(conditions)
        return (f"EVALUATE\nCALCULATETABLE(\n    TOPN({row_limit}, '{safe}'),\n"
                f"    {cond_str}\n)")
    return f"EVALUATE TOPN({row_limit}, '{safe}')"


@app.route('/api/export/count', methods=['POST'])
def export_count():
    """Return the number of rows that match the current filters (used to warn before capping at 3,000)."""
    body           = request.get_json(silent=True) or {}
    table          = (body.get('table_override') or '').strip() or None
    report_filters = body.get('report_filters') or []
    slicer_states  = body.get('slicer_states')  or []

    try:
        access_token = get_access_token()
        dataset_id   = _get_dataset_id(access_token)
        headers      = {'Authorization': f'Bearer {access_token}',
                        'Content-Type': 'application/json'}

        if not table:
            table = _extract_table_name(report_filters, slicer_states)
        if not table:
            return jsonify({'count': None}), 200

        safe = table.replace("'", "''")
        conditions = []
        for f in (report_filters or []):
            c = _filter_to_dax(f)
            if c: conditions.append(c)
        for s in (slicer_states or []):
            for sf in ((s.get('state') or {}).get('filters') or []):
                c = _filter_to_dax(sf)
                if c: conditions.append(c)

        if conditions:
            cond_str = ', '.join(conditions)
            dax = f"EVALUATE ROW(\"count\", COUNTROWS(CALCULATETABLE('{safe}', {cond_str})))"
        else:
            dax = f"EVALUATE ROW(\"count\", COUNTROWS('{safe}'))"

        url = f"https://api.powerbi.com/v1.0/myorg/datasets/{dataset_id}/executeQueries"
        qr  = requests.post(url, headers=headers, json={"queries": [{"query": dax}]})
        if qr.status_code != 200:
            return jsonify({'count': None}), 200

        rows = (qr.json().get('results') or [{}])[0].get('tables', [{}])[0].get('rows', [])
        count = list(rows[0].values())[0] if rows else None
        return jsonify({'count': count})
    except Exception:
        return jsonify({'count': None}), 200


@app.route('/api/export/data', methods=['GET', 'POST'])
def export_data():
    """Export a dataset table as CSV or Excel with current Power BI filters applied (max 3,000 rows)."""
    if request.method == 'POST':
        body           = request.get_json(silent=True) or {}
        table          = (body.get('table_override') or '').strip() or None
        report_filters = body.get('report_filters') or []
        slicer_states  = body.get('slicer_states')  or []
        visual_title   = (body.get('visual_title') or '').strip()
        fmt            = (body.get('fmt') or 'csv').strip().lower()
    else:
        table          = (request.args.get('table') or '').strip() or None
        report_filters = []
        slicer_states  = []
        visual_title   = table or 'export'
        fmt            = (request.args.get('fmt') or 'csv').strip().lower()

    if fmt not in ('csv', 'xlsx'):
        fmt = 'csv'

    try:
        access_token = get_access_token()
        dataset_id   = _get_dataset_id(access_token)
        headers      = {'Authorization': f'Bearer {access_token}',
                        'Content-Type': 'application/json'}

        if not table:
            table = _extract_table_name(report_filters, slicer_states)
        if not table:
            return jsonify({
                'success': False,
                'error': ('Could not identify the dataset table. '
                          'Apply at least one slicer then try again, '
                          'or use the manual table-name input.')
            }), 400

        dax = _build_dax(table, report_filters, slicer_states)
        url = f"https://api.powerbi.com/v1.0/myorg/datasets/{dataset_id}/executeQueries"
        qr  = requests.post(url, headers=headers,
                            json={"queries": [{"query": dax}],
                                  "serializerSettings": {"includeNulls": True}})
        if qr.status_code != 200:
            return jsonify({'success': False,
                            'error': f'Query failed ({qr.status_code}): {qr.text[:400]}'}), 500

        rows = (qr.json().get('results') or [{}])[0].get('tables', [{}])[0].get('rows', [])
        if not rows:
            return jsonify({'success': False, 'error': 'No data returned.'}), 404

        def clean_col(name):
            return re.sub(r'^[^[]*\[', '', name).rstrip(']')

        raw_cols   = list(rows[0].keys())
        clean_cols = [clean_col(c) for c in raw_cols]
        base_name  = re.sub(r'[^\w\s-]', '', visual_title or table).strip().replace(' ', '_')

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = (visual_title or table)[:31]
            ws.append(clean_cols)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append([row.get(c, '') for c in raw_cols])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = base_name + '.xlsx'
            return Response(
                buf.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{fname}"'}
            )
        else:
            buf = io.StringIO()
            w   = csv.writer(buf)
            w.writerow(clean_cols)
            for row in rows:
                w.writerow([row.get(c, '') for c in raw_cols])
            fname = base_name + '.csv'
            return Response(buf.getvalue(), mimetype='text/csv',
                            headers={'Content-Disposition': f'attachment; filename="{fname}"'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/convert', methods=['POST'])
def convert_csv_to_xlsx():
    """Accept a CSV string from the browser (visual.exportData result) and return it as Excel."""
    body     = request.get_json(silent=True) or {}
    csv_data = body.get('csv_data', '')
    filename = re.sub(r'[^\w\s-]', '', body.get('filename', 'export')).strip().replace(' ', '_') or 'export'

    if not csv_data:
        return jsonify({'error': 'No CSV data provided.'}), 400

    try:
        reader = csv.reader(io.StringIO(csv_data))
        rows   = list(reader)
        if not rows:
            return jsonify({'error': 'CSV is empty.'}), 400

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = filename[:31]
        for i, row in enumerate(rows):
            ws.append(row)
            if i == 0:
                for cell in ws[1]:
                    cell.font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}.xlsx"'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("=" * 50)
    print("Power BI Embedded Demo Starting...")
    print("=" * 50)
    print(f"Main App: http://localhost:8080/")
    print(f"Test API: http://localhost:8080/test")
    print(f"Health:   http://localhost:8080/health")
    print("=" * 50)
    print("Default Config:")
    print(f"   Tenant:  {DEFAULT_CONFIG['tenant_id']}")
    print(f"   Group:   {DEFAULT_CONFIG['group_id']}")
    print(f"   Report:  {DEFAULT_CONFIG['report_id']}")
    print("=" * 50)

    app.run(host='0.0.0.0', port=8080, debug=True)
